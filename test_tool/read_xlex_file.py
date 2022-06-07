from openpyxl import load_workbook


class ReadData:

    @staticmethod
    def read_data(file_name, sheet_name):
        x = load_workbook(file_name)
        w = x[sheet_name]
        max_r = w.max_row
        max_c = w.max_column
        print("最大行是：", max_r, "最大列：", max_c)
        data = [{w.cell(1, j).value: w.cell(i, j).value for j in range(1, max_c + 1)} for i in range(2, max_r + 1)]
        return data


if __name__ == '__main__':
    d = ReadData().read_data(r"/data/test_data.xlsx", "Sheet1")
    print(d)
